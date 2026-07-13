"""Exploración rápida de las bases reales (carpeta `data/01_raw/`).

Lee en modo streaming (openpyxl read_only) para no cargar en memoria la hoja
gigante de la Alcaldía. Imprime cabeceras, una muestra de filas y la
distribución de las columnas clave para diseñar el ingest/entrenamiento.

Uso:  python -m ml.explore
"""
from __future__ import annotations

from collections import Counter

from openpyxl import load_workbook

from app.config import RAW_DATA_DIR

DB_DIR = RAW_DATA_DIR
SAMPLE = 8000


def find(pattern: str):
    for p in DB_DIR.glob("*.xlsx"):
        if pattern.lower() in p.name.lower():
            return p
    raise FileNotFoundError(pattern)


def sheet_named(wb, needle: str):
    for name in wb.sheetnames:
        if needle.lower() in name.lower():
            return name
    return None


def explore_sheet(path, needle, key_cols, sample=SAMPLE):
    wb = load_workbook(path, read_only=True, data_only=True)
    name = sheet_named(wb, needle)
    print(f"\n===== {path.name} → hoja «{name}» =====")
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    print("Columnas:", header)
    idx = {h: i for i, h in enumerate(header)}

    counters = {c: Counter() for c in key_cols if c in idx}
    n = 0
    sample_rows = []
    for r in rows:
        n += 1
        if len(sample_rows) < 8:
            sample_rows.append(r)
        if n <= sample:
            for c in counters:
                counters[c][r[idx[c]]] += 1
        if n >= sample and len(sample_rows) >= 8:
            # seguimos contando filas totales sin acumular
            pass
    print(f"Filas (excl. cabecera): {n}")
    print("Muestra:")
    for sr in sample_rows[:6]:
        print("  ", sr)
    for c, cnt in counters.items():
        top = cnt.most_common(12)
        print(f"\n-- {c} -- ({len(cnt)} distintos) top:")
        for v, k in top:
            print(f"    {v!r}: {k}")
    wb.close()


def main():
    # Alcaldía: la base de hurtos consolidada (fecha + hora + comuna + barrio)
    explore_sheet(
        find("consolidado"), "alca sec",
        key_cols=["FECHA_HECHO", "HORA_HECHO", "TIPO_SITIO", "ARMA_EMPLEADA",
                  "COMUNA", "BARRIO", "CANTIDAD", "SEXO", "EDAD"],
    )
    # Policía: agregado (sin hora)
    explore_sheet(
        find("homologado"), "polc",
        key_cols=["Temática", "Año", "Mes", "Día", "Zona", "Clase de Sitio", "Cantidad"],
    )
    # Ubicación de CAI/cuadrantes
    wb = load_workbook(find("polic"), read_only=True, data_only=True)
    name = sheet_named(wb, "limpio")
    print(f"\n===== ubicaciones → hoja «{name}» =====")
    ws = wb[name]
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i < 6:
            print("  ", r)
        else:
            break
    wb.close()


if __name__ == "__main__":
    main()

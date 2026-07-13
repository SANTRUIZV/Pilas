"""Ingesta de las bases reales (`data/01_raw/`) → CSVs limpios de entrenamiento.

Fuente principal: hoja «ALCA SEC UNIDOS 2010-2026» de
`Consolidado_secretaria_alcaldia_2010_2026.xlsx` — la base de HURTOS de la
Secretaría/Alcaldía (fusión de 2010-2019 y 2019-2026), ~219k incidentes de Cali
con fecha, hora, comuna y barrio. Se agrega CANTIDAD a la malla
(año, comuna, hora, día_semana, mes) que alimenta el modelo de riesgo. El año,
el mes y el día de la semana se derivan de FECHA_HECHO (que viene como datetime o
como texto `dd/mm/aaaa`).

Salidas en `data/03_primary/`:
  - incidents_cali.csv     año, comuna, hour, weekday, month, is_holiday, count  (modelo de riesgo)
  - crime_monthly.csv      conflictividad, year, month, count        (tendencias; aquí solo «Hurto persona»)
  - comuna_totals.csv      comuna, count                             (tabla de comunas)
  - cai_locations.csv      name, lat, lon, phone, address            (unidades de Policía)
  - health_services.csv    name, lat, lon, phone, address            (servicios de salud)

Uso:  python -m ml.ingest
"""
from __future__ import annotations

import csv
import heapq
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime, time

from openpyxl import load_workbook

from app.config import DATA_DIR, RAW_DATA_DIR
from app.holidays import is_holiday

DB_DIR = RAW_DATA_DIR


def _norm(s) -> str:
    s = str(s).strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def _find(pattern: str):
    for p in DB_DIR.glob("*.xlsx"):
        if pattern.lower() in p.name.lower():
            return p
    raise FileNotFoundError(f"No se encontró xlsx con '{pattern}' en {DB_DIR}")


def _sheet(wb, needle: str):
    for name in wb.sheetnames:
        if needle.lower() in _norm(name):
            return name
    raise KeyError(needle)


def _to_hour(v) -> int | None:
    if isinstance(v, time):
        return v.hour
    if isinstance(v, datetime):
        return v.hour
    if isinstance(v, str) and ":" in v:
        try:
            return int(v.split(":")[0])
        except ValueError:
            return None
    return None


def _to_date(v) -> date | None:
    """FECHA_HECHO normalizada a `date`. Acepta datetime/date de Excel o texto
    en formato colombiano `dd/mm/aaaa` (la base trae ambos)."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _to_comuna(v) -> int | None:
    try:
        c = int(v)
    except (TypeError, ValueError):
        return None
    return c if 1 <= c <= 22 else None


# La base consolidada es exclusivamente de hurtos → toda fila cuenta como
# «Hurto persona» para crime_monthly (id de delito «hurto-personas» en el front).
_HURTO_CONFLICTIVIDAD = "Hurto persona"


def ingest_alcaldia():
    path = _find("consolidado")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[_sheet(wb, "alca sec")]   # «ALCA SEC UNIDOS 2010-2026»
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    ix = {h: i for i, h in enumerate(header)}

    grid = defaultdict(int)            # (year, comuna, hour, weekday, month, holiday) -> count
    crime_monthly = defaultdict(int)   # (conflictividad, year, month) -> count
    comuna_totals = Counter()          # comuna -> count
    seen = kept = 0

    for r in rows:
        seen += 1
        comuna = _to_comuna(r[ix["COMUNA"]])
        hour = _to_hour(r[ix["HORA_HECHO"]])
        if comuna is None or hour is None:
            continue
        d = _to_date(r[ix["FECHA_HECHO"]])
        if d is None:
            continue
        year, month, weekday = d.year, d.month, d.weekday()
        qty = r[ix["CANTIDAD"]]
        qty = int(qty) if isinstance(qty, (int, float)) else 1

        holiday = 1 if is_holiday(d) else 0
        grid[(year, comuna, hour, weekday, month, holiday)] += qty
        comuna_totals[comuna] += qty
        crime_monthly[(_HURTO_CONFLICTIVIDAD, year, month)] += qty
        kept += 1

    wb.close()
    print(f"· Alcaldía (hurtos 2010–2026): {seen:,} filas leídas, {kept:,} válidas, {len(grid):,} celdas de malla")

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    with open(DATA_DIR / "incidents_cali.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["year", "comuna", "hour", "weekday", "month", "is_holiday", "count"])
        for (year, comuna, hour, weekday, month, holiday), c in sorted(grid.items()):
            w.writerow([year, comuna, hour, weekday, month, holiday, c])

    with open(DATA_DIR / "crime_monthly.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["conflictividad", "year", "month", "count"])
        for (conf, year, month), c in sorted(crime_monthly.items()):
            w.writerow([conf, year, month, c])

    with open(DATA_DIR / "comuna_totals.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["comuna", "count"])
        for comuna, c in sorted(comuna_totals.items()):
            w.writerow([comuna, c])


# ── Estadísticas ciudadanas (modalidad, sitio, víctima, barrio) ──────────────
# La base fusiona dos fuentes con taxonomías distintas («ARMA DE FUEGO» vs «Arma
# de fuego», etc.); estas funciones normalizan a categorías canónicas limpias.

def _modalidad(v) -> str:
    """ARMA_EMPLEADA → modalidad canónica del hurto."""
    s = _norm(v)
    if not s or s in ("none", "sin dato"):
        return "Sin dato"
    if "fuego" in s:
        return "Arma de fuego"
    if "blanca" in s or "cortante" in s or "punzante" in s:
        return "Arma blanca"
    if "contundente" in s:
        return "Objeto contundente"
    if "toxica" in s or "escopolamina" in s or "sustancia" in s:
        return "Escopolamina / sustancia"
    if "sin arma" in s or "sin empleo" in s:
        return "Sin arma (atraco directo)"
    return "Sin dato"


def _sitio(v) -> str:
    """TIPO_SITIO → tipo de lugar canónico. El campo es muy ruidoso (trae
    direcciones libres), así que lo no reconocible cae a «Otro / sin clasificar»."""
    s = _norm(v)
    if not s or s in ("none", "sin dato"):
        return "Otro / sin clasificar"
    if ("via" in s and "publ" in s) or "semaforo" in s:
        return "Vía pública"
    if "residencia" in s or "vivienda" in s or "apartament" in s or "conjunto" in s:
        return "Residencia"
    if "vehiculo" in s or "automotor" in s:
        return "Interior de vehículo"
    if "mio" in s or "masivo" in s or "transporte" in s or s.startswith("bus"):
        return "Transporte público"
    if "comercial" in s or "local" in s or "tienda" in s or "almacen" in s:
        return "Comercio"
    if "cajero" in s or "banco" in s:
        return "Cajero / banco"
    if "parqueadero" in s:
        return "Parqueadero"
    return "Otro / sin clasificar"


def _sexo(v) -> str:
    s = _norm(v)
    if "hombre" in s or s == "m" or "masculino" in s:
        return "Hombre"
    if "mujer" in s or "femenino" in s:
        return "Mujer"
    return "Sin dato"


_EDAD_BANDS = ["< 18", "18-25", "26-35", "36-45", "46-60", "60+"]


def _edad_band(e) -> str | None:
    if not isinstance(e, (int, float)) or not (0 < e < 120):
        return None
    e = int(e)
    if e < 18:
        return "< 18"
    if e <= 25:
        return "18-25"
    if e <= 35:
        return "26-35"
    if e <= 45:
        return "36-45"
    if e <= 60:
        return "46-60"
    return "60+"


def ingest_hurto_stats():
    """Agrega dimensiones cualitativas de la base de hurtos para la vista de
    «Estadísticas» de la app ciudadana: modalidad (arma), tipo de sitio, perfil
    de la víctima (sexo, edad) y barrios más afectados. Una pasada por la hoja.

    Salidas en `data/03_primary/`:
      - stats_modalidad.csv   modalidad, count
      - stats_sitio.csv       sitio, count
      - stats_sexo.csv        sexo, count
      - stats_edad.csv        band, count
      - stats_barrio.csv      barrio, comuna, count   (TODOS los barrios de la base)
      - recent_reports.csv    date, hour, comuna, barrio, modalidad, sitio  (50 más recientes)
    """
    path = _find("consolidado")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[_sheet(wb, "alca sec")]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    ix = {h: i for i, h in enumerate(header)}

    modalidad = Counter()
    sitio = Counter()
    sexo = Counter()
    edad = Counter()
    # Barrios agrupados por nombre NORMALIZADO (minúsculas, sin tildes) para que
    # «SAN PEDRO», «San Pedro» y «san pedro» cuenten como uno; se conserva la
    # grafía más frecuente como nombre visible.
    barrio = Counter()                                       # norm → count
    barrio_display: dict[str, Counter] = defaultdict(Counter)  # norm → grafías
    barrio_comuna: dict[str, Counter] = defaultdict(Counter)   # norm → comunas
    recent: list = []        # min-heap (sortkey, seq, record): 50 reportes más recientes
    seq = 0
    RECENT_LIMIT = 50

    for r in rows:
        qty = r[ix["CANTIDAD"]]
        qty = int(qty) if isinstance(qty, (int, float)) else 1
        mod = _modalidad(r[ix["ARMA_EMPLEADA"]])
        sit = _sitio(r[ix["TIPO_SITIO"]])
        modalidad[mod] += qty
        sitio[sit] += qty
        sexo[_sexo(r[ix["SEXO"]])] += qty
        band = _edad_band(r[ix["EDAD"]])
        if band:
            edad[band] += qty
        b = r[ix["BARRIO"]]
        bn = str(b).strip() if b is not None else ""
        comuna = _to_comuna(r[ix["COMUNA"]])
        bkey = _norm(bn)
        if bn and bkey not in ("sin dato", "none", "rural"):
            barrio[bkey] += qty
            barrio_display[bkey][bn] += qty
            if comuna is not None:
                barrio_comuna[bkey][comuna] += qty

        # Reportes recientes: conserva los 50 con fecha+hora más recientes.
        d = _to_date(r[ix["FECHA_HECHO"]])
        hr = _to_hour(r[ix["HORA_HECHO"]])
        if d is not None and hr is not None:
            key = d.toordinal() * 24 + hr
            rec = (d.isoformat(), hr, comuna if comuna is not None else "", bn, mod, sit)
            if len(recent) < RECENT_LIMIT:
                heapq.heappush(recent, (key, seq, rec))
                seq += 1
            elif key > recent[0][0]:
                heapq.heappushpop(recent, (key, seq, rec))
                seq += 1
    wb.close()

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    def _write_counter(name, col, counter, order=None):
        items = [(k, counter[k]) for k in order if k in counter] if order \
            else counter.most_common()
        with open(DATA_DIR / name, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow([col, "count"])
            for k, c in items:
                w.writerow([k, c])

    _write_counter("stats_modalidad.csv", "modalidad", modalidad)
    _write_counter("stats_sitio.csv", "sitio", sitio)
    _write_counter("stats_sexo.csv", "sexo", sexo)
    _write_counter("stats_edad.csv", "band", edad, order=_EDAD_BANDS)

    with open(DATA_DIR / "stats_barrio.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["barrio", "comuna", "count"])
        for bkey, c in barrio.most_common():
            display = barrio_display[bkey].most_common(1)[0][0]
            cm = barrio_comuna[bkey].most_common(1)
            w.writerow([display, cm[0][0] if cm else "", c])

    with open(DATA_DIR / "recent_reports.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["date", "hour", "comuna", "barrio", "modalidad", "sitio"])
        for _key, _seq, rec in sorted(recent, key=lambda x: x[0], reverse=True):
            w.writerow(list(rec))

    print(f"· Estadísticas hurto: modalidad={len(modalidad)} · sitio={len(sitio)} · "
          f"barrios={len(barrio)} · recientes={len(recent)}")


def _parse_coords(lat_v, lon_v):
    """Coordenadas en dos celdas numéricas (a veces sin punto decimal: 34191 →
    3.4191) o combinadas como «lat, lon» en una sola celda (hoja «Limpio»).
    Devuelve (lat, lon) validado contra el área de Cali, o None."""
    if isinstance(lat_v, str) and "," in lat_v:
        parts = lat_v.split(",")
        try:
            lat, lon = float(parts[0]), float(parts[1])
        except (IndexError, ValueError):
            return None
    else:
        try:
            lat, lon = float(lat_v), float(lon_v)
        except (TypeError, ValueError):
            return None
    if abs(lat) > 90:
        lat /= 10000.0
    if abs(lon) > 180:
        lon /= 10000.0
    if not (3.2 <= lat <= 3.6 and -77.2 <= lon <= -76.2):
        return None
    return round(lat, 5), round(lon, 5)


def _phone(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        return str(int(v))
    s = str(v).strip()
    return "" if s.upper() in {"#N/A", "N/A", "NA", "NONE"} else s


def _read_units(ws):
    """Lee una hoja geolocalizada de unidades → [{name, lat, lon, phone, address}]."""
    rows = ws.iter_rows(values_only=True)
    header = [(_norm(h) if h is not None else "") for h in next(rows)]

    def col(*names):
        for n in names:
            for i, h in enumerate(header):
                if n in h:
                    return i
        return None

    i_name = col("nombre")
    i_lat, i_lon = col("latitud"), col("longitud")
    i_tel = col("telefono")
    i_addr = col("direcci", "referencia")  # "direcci" tolera mojibake en la cabecera
    out, skipped = [], 0
    if i_name is None or i_lat is None:
        return out, skipped
    for r in rows:
        if not any(c is not None for c in r):
            continue
        coords = _parse_coords(r[i_lat], r[i_lon] if i_lon is not None else None)
        if coords is None:
            skipped += 1
            continue
        addr = str(r[i_addr]).strip() if i_addr is not None and r[i_addr] is not None else ""
        if _parse_coords(addr, None):  # dirección rellenada con coordenadas → descartarla
            addr = ""
        out.append({
            "name": str(r[i_name]).strip() if r[i_name] is not None else "",
            "lat": coords[0], "lon": coords[1],
            "phone": _phone(r[i_tel]) if i_tel is not None else "",
            "address": addr,
        })
    return out, skipped


def ingest_cai():
    """«Hoja3» es el set geolocalizado más completo (CAI + estaciones +
    subestaciones, con dirección); la hoja «Limpio» trae coordenadas verificadas
    (más precisas) y unidades nuevas. Partimos de Hoja3 y sobreescribimos /
    añadimos con «Limpio», casando por nombre normalizado."""
    path = _find("polic")
    wb = load_workbook(path, read_only=True, data_only=True)

    def key(name: str) -> str:
        return " ".join(_norm(name).split())

    units: dict[str, dict] = {}
    order: list[str] = []
    skipped = 0
    found = []
    for cand in ("hoja3", "limpio"):
        try:
            sheet = _sheet(wb, cand)
        except KeyError:
            continue
        found.append(sheet)
        rows, sk = _read_units(wb[sheet])
        skipped += sk
        for u in rows:
            k = key(u["name"])
            if k in units:
                prev = units[k]
                prev["lat"], prev["lon"] = u["lat"], u["lon"]
                if u["phone"]:
                    prev["phone"] = u["phone"]
                if u["address"]:
                    prev["address"] = u["address"]
            else:
                units[k] = u
                order.append(k)
    wb.close()
    if not found:
        print("· Ubicaciones: hoja geolocalizada no encontrada; omitido")
        return

    with open(DATA_DIR / "cai_locations.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["name", "lat", "lon", "phone", "address"])
        for k in order:
            u = units[k]
            w.writerow([u["name"], u["lat"], u["lon"], u["phone"], u["address"]])
    print(f"· Unidades de Policía (hojas «{'» + «'.join(found)}»): {len(order):,} geolocalizadas ({skipped} descartadas)")


# Tokens que se conservan en mayúsculas al embellecer nombres de prestadores.
_KEEP_UPPER = {"IPS", "EPS", "ESE", "E.S.E.", "E.S.E", "SAS", "S.A.S", "S.A.S.", "S.A",
               "S.A.", "S.A.M.U.", "LTDA", "UBA", "AIP", "CEM", "II", "III"}
_LOWER = {"de", "del", "y", "al", "con", "en", "a"}
_ACCENTS = {"clinica": "Clínica", "fundacion": "Fundación", "medico": "Médico",
            "medica": "Médica", "atencion": "Atención", "psiquiatrico": "Psiquiátrico"}


def _pretty(name: str) -> str:
    """«FUNDACION VALLE DEL LILI» → «Fundación Valle del Lili»."""
    out = []
    for w in str(name).split():
        if w.upper() in _KEEP_UPPER:
            out.append(w.upper())
        elif w.lower() in _ACCENTS:
            out.append(_ACCENTS[w.lower()])
        elif w.lower() in _LOWER and out:
            out.append(w.lower())
        else:
            out.append(w.capitalize())
    return " ".join(out)


def ingest_salud():
    """Servicios de salud habilitados con urgencias en Cali (hoja «LIMPIO» del
    Excel de servicios habilitados): nombre, coordenadas, teléfono y dirección."""
    try:
        path = _find("salud")
    except FileNotFoundError:
        print("· Salud: xlsx de servicios habilitados no encontrado; omitido")
        return
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[_sheet(wb, "limpio")]
    rows = ws.iter_rows(values_only=True)
    header = [(_norm(h) if h is not None else "") for h in next(rows)]

    def col(*names):
        for n in names:
            for i, h in enumerate(header):
                if n in h:
                    return i
        return None

    i_name = col("servicio", "nombre")
    i_lat, i_lon = col("latitud"), col("longitud")
    i_tel = col("telefono")
    i_addr = col("direcci", "columna")  # la dirección quedó en «Columna1»
    out, skipped = [], 0
    for r in rows:
        if not any(c is not None for c in r):
            continue
        coords = _parse_coords(r[i_lat], r[i_lon] if i_lon is not None else None)
        if coords is None:
            skipped += 1
            continue
        out.append([
            _pretty(r[i_name]) if r[i_name] is not None else "",
            coords[0], coords[1],
            _phone(r[i_tel]) if i_tel is not None else "",
            str(r[i_addr]).strip() if i_addr is not None and r[i_addr] is not None else "",
        ])
    wb.close()

    with open(DATA_DIR / "health_services.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["name", "lat", "lon", "phone", "address"])
        w.writerows(out)
    print(f"· Servicios de salud (hoja «LIMPIO»): {len(out):,} geolocalizados ({skipped} descartados)")


def ingest_cuadrantes():
    """Directorio de cuadrantes de Cali (hoja «ORIGINAL»): estación, CAI, código,
    cuadrante y teléfono. No tiene coordenadas → es un listado consultable, no
    va al mapa."""
    path = _find("polic")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[_sheet(wb, "original")]
    except KeyError:
        wb.close()
        print("· Cuadrantes: hoja 'ORIGINAL' no encontrada; omitido")
        return
    rows = ws.iter_rows(values_only=True)
    header = [(_norm(h) if h is not None else "") for h in next(rows)]

    def col(*names):
        for n in names:
            for i, h in enumerate(header):
                if n in h:
                    return i
        return None

    def colexact(name):
        for i, h in enumerate(header):
            if h == name:
                return i
        return None

    i_ciudad = col("ciudad", "municipio")
    i_unidad = colexact("unidad")
    i_cai = col("tipo")
    i_codigo = col("codigo")
    i_cuad = colexact("cuadrante")
    i_phone = col("celular", "telefono", "numero")

    def s(r, i):
        return str(r[i]).strip() if i is not None and r[i] is not None else ""

    def phone(r, i):
        v = r[i] if i is not None else None
        if v is None:
            return ""
        if isinstance(v, float):
            return str(int(v))
        return str(v).strip()

    out, total = [], 0
    for r in rows:
        if i_ciudad is None:
            break
        if _norm(r[i_ciudad]) != "cali":
            continue
        total += 1
        cai = s(r, i_cai)
        if "error" in _norm(cai):
            cai = ""
        out.append([s(r, i_unidad), cai, s(r, i_codigo), s(r, i_cuad), phone(r, i_phone)])
    wb.close()

    out.sort(key=lambda x: (x[0], x[3]))
    with open(DATA_DIR / "cuadrantes_cali.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["estacion", "cai", "codigo", "cuadrante", "phone"])
        w.writerows(out)
    print(f"· Cuadrantes de Cali (hoja «ORIGINAL»): {len(out):,} registros")


# ── Violencia intrafamiliar (MinDefensa, corte Cali) ─────────────────────────
_CALI_COD_MUNI = 76001


def ingest_violencia_intrafamiliar():
    """Base nacional de violencia intrafamiliar de MinDefensa (hoja «DATOS»,
    ~657k filas país): conteos diarios por municipio. Se filtra Cali
    (COD_MUNI 76001) y se agrega CANTIDAD por año × mes.

    Salida: `vif_monthly.csv` (year, month, count) — alimenta la serie
    «Violencia intrafamiliar» de /gov/series, las alertas y el briefing."""
    try:
        path = _find("intrafamiliar")
    except FileNotFoundError:
        print("· VIF: xlsx de MinDefensa no encontrado; omitido")
        return
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[_sheet(wb, "datos")]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    ix = {str(h).strip(): i for i, h in enumerate(header)}

    monthly = defaultdict(int)   # (year, month) -> count
    seen = kept = 0
    for r in rows:
        seen += 1
        try:
            if int(r[ix["COD_MUNI"]]) != _CALI_COD_MUNI:
                continue
        except (TypeError, ValueError):
            continue
        d = _to_date(r[ix["FECHA HECHO"]])
        if d is None:
            continue
        qty = r[ix["CANTIDAD"]]
        qty = int(qty) if isinstance(qty, (int, float)) else 1
        monthly[(d.year, d.month)] += qty
        kept += 1
    wb.close()

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(DATA_DIR / "vif_monthly.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["year", "month", "count"])
        for (y, m), c in sorted(monthly.items()):
            w.writerow([y, m, c])
    years = sorted({y for y, _ in monthly})
    total = sum(monthly.values())
    print(f"· Violencia intrafamiliar (MinDefensa · Cali): {seen:,} filas país, "
          f"{kept:,} de Cali, {total:,} casos ({years[0]}–{years[-1]})" if years else
          "· Violencia intrafamiliar: sin filas de Cali")


# ── Violencia de género en Cali 2013–2022 (Datos Abiertos) ───────────────────
# El xlsx trae 4 hojas con esquemas distintos (2013-2018, 2019, 2020, 2021-2022);
# estas funciones normalizan cada dimensión a categorías canónicas comunes.

def _gv_tipo(v) -> str | None:
    """tipo_violenc → tipo canónico. Devuelve None para ruido (la hoja 2019 trae
    algunas filas con valores de otra columna, p. ej. «Estudiante»)."""
    s = _norm(v)
    if not s or s in ("none", "sin dato", "ninguna"):
        return None
    if "fisic" in s:
        return "Física"
    if "psicolog" in s:
        return "Psicológica"
    if "negligencia" in s or "privacion" in s or "abandono" in s:
        return "Negligencia y abandono"
    if any(k in s for k in ("sexual", "violacion", "trata", "acoso", "turismo")):
        return "Sexual"
    return None


def _gv_sexo(v) -> str:
    s = _norm(v)
    if "hombre" in s or "masculino" in s or s == "m":
        return "Hombre"
    if "mujer" in s or "femenino" in s or s == "f":
        return "Mujer"
    return "Sin dato"


def _gv_comuna(v) -> int | None:
    """comuna en formatos mixtos: 13 · «Comuna 21» · «comuna 15 » · «Sin dato».
    Valores >22 (corregimientos) se descartan del ranking por comuna."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        c = int(v)
    else:
        s = _norm(v).replace("comuna", "").strip()
        if not s.isdigit():
            return None
        c = int(s)
    return c if 1 <= c <= 22 else None


def _gv_agresor(v) -> str:
    """relacion_fam_victima (hojas 2019 y 2020) → categoría canónica."""
    s = _norm(v)
    if "ex" in s and "pareja" in s:
        return "Ex-pareja"
    if "pareja" in s:
        return "Pareja"
    if s in ("familiar", "madre", "padre"):
        return "Otro familiar"
    if s == "ninguno":
        return "No familiar"
    return "Sin dato"


def ingest_violencia_genero():
    """Eventos de violencia de género en Cali 2013–2022 (Datos Abiertos, ~63k
    eventos individuales). Agrega por año, comuna, tipo de violencia, sexo y
    edad de la víctima, y relación con el agresor.

    Salidas en `data/03_primary/` (todas `label,count` salvo la anual):
      - gv_yearly.csv    year, count                (todas las hojas)
      - gv_comuna.csv    comuna, count              (todas las hojas)
      - gv_tipo.csv      tipo, count                (hojas 2013-2020; 2021-2022
                                                     solo distingue sexual/no sexual)
      - gv_sexo.csv      sexo, count                (todas las hojas)
      - gv_edad.csv      band, count                (todas las hojas)
      - gv_agresor.csv   agresor, count             (hojas 2019-2020)
    """
    try:
        path = _find("violencia-de-genero")
    except FileNotFoundError:
        print("· Violencia de género: xlsx no encontrado; omitido")
        return
    wb = load_workbook(path, read_only=True, data_only=True)

    yearly = Counter()
    comuna = Counter()
    tipo = Counter()
    sexo = Counter()
    edad = Counter()
    agresor = Counter()

    def process(sheet: str, year_col: str | None, fixed_year: int | None,
                sexo_col: str, edad_col: str, comuna_col: str,
                tipo_col: str | None, agresor_col: str | None):
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        ix = {str(h).strip(): i for i, h in enumerate(next(rows))}
        n = 0
        for r in rows:
            if not any(c is not None for c in r):
                continue
            n += 1
            y = fixed_year
            if year_col is not None:
                try:
                    y = int(r[ix[year_col]])
                except (TypeError, ValueError):
                    y = None
            if y is not None:
                yearly[y] += 1
            c = _gv_comuna(r[ix[comuna_col]])
            if c is not None:
                comuna[c] += 1
            sexo[_gv_sexo(r[ix[sexo_col]])] += 1
            band = _edad_band(r[ix[edad_col]])
            if band:
                edad[band] += 1
            if tipo_col is not None:
                t = _gv_tipo(r[ix[tipo_col]])
                if t:
                    tipo[t] += 1
            if agresor_col is not None:
                agresor[_gv_agresor(r[ix[agresor_col]])] += 1
        return n

    n1 = process("2013-2018", "ano", None, "sexo", "edad", "comuna",
                 "tipo_violenc", None)
    n2 = process("2019", None, 2019, "sexo_vict", "edad", "comuna",
                 "tipo_violenc", "relacion_fam_victima")
    n3 = process("2020", None, 2020, "sexo_vict", "edad", "comuna",
                 "tipo_violenc", "relacion_fam_victima")
    n4 = process("2021-2022", "ano_", None, "sexo_", "edad_", "com_",
                 None, None)
    wb.close()

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    def _write(name, col, items):
        with open(DATA_DIR / name, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow([col, "count"])
            w.writerows(items)

    _write("gv_yearly.csv", "year", sorted(yearly.items()))
    _write("gv_comuna.csv", "comuna", sorted(comuna.items(), key=lambda x: -x[1]))
    _write("gv_tipo.csv", "tipo", tipo.most_common())
    _write("gv_sexo.csv", "sexo", sexo.most_common())
    _write("gv_edad.csv", "band", [(b, edad[b]) for b in _EDAD_BANDS if b in edad])
    _write("gv_agresor.csv", "agresor", agresor.most_common())

    total = n1 + n2 + n3 + n4
    print(f"· Violencia de género (Cali 2013–2022): {total:,} eventos "
          f"({n1:,} + {n2:,} + {n3:,} + {n4:,}) · comunas={len(comuna)} · tipos={len(tipo)}")


def main():
    # La consola de Windows usa cp1252 por defecto y no puede imprimir «✓»/«·».
    try:
        import sys
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    ingest_alcaldia()
    ingest_hurto_stats()
    ingest_cai()
    ingest_cuadrantes()
    ingest_salud()
    ingest_violencia_intrafamiliar()
    ingest_violencia_genero()
    print(f"✓ CSVs en {DATA_DIR}")


if __name__ == "__main__":
    main()
